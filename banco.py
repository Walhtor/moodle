# Script para la carga de preguntas desde una planilla excel en un banco de preguntas moodle

# import openpyxl module
import openpyxl

path = "/home/sysadmin/PycharmProjects/moodle/banco-preguntas3.xlsx"

# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path)

# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active

# Cell objects also have row, column,
# and coordinate attributes that provide
# location information for the cell.

# Note: The first row or
# column integer is 1, not 0.
cell_obj = sheet_obj.cell(row=1, column=2)
# Cell object is created by using
# sheet object's cell() method.
f = open('banco_preguntas.txt', 'w+')

# sheet = wb_obj.get_sheet_by_name('Hoja1')
#print(sheet_obj.max_row)
max_fila = sheet_obj.max_row
#print(sheet_obj.max_column)
max_columna = sheet_obj.max_column

#print(tuple(sheet_obj['A1': 'B' + str(max_fila)]))
f.write(cell_obj.value)
f.close()

for filas in sheet_obj['A1':'B' + str(max_fila)]:
    for celda in filas:
            #tipo = ""
            if type(celda.value) == int and celda.value > 0:
                #print("Pregunta")
                tipo = "Pregunta"
            elif type(celda.value) == int and celda.value < 0:
                tipo = "Correcta"
            elif type(celda.value) == int and celda.value == 0:
                tipo = "Respuesta"
            else:
                if tipo == "Pregunta":
                    print()
                    print("Pregunta", celda.value + " {")
                elif tipo == "Correcta":
                    print("Correcta","= " + celda.value)
                else:
                    print("Respuesta","~ " + celda.value)
    #print('--- END OF ROW ---')

cell_obj = sheet_obj.cell(row=1, column=2)

# Print value of cell object
# using the value attribute
#print(cell_obj.value)
